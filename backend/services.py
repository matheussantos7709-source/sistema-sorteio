from datetime import datetime, timedelta
from typing import Dict, Any, List, Tuple, Optional
from io import BytesIO
import unicodedata
import re

from openpyxl import Workbook, load_workbook

try:
    from backend.database import conectar, registrar_vencedor
except Exception:
    from database import conectar, registrar_vencedor


# -----------------------------
# HELPERS
# -----------------------------

def _norm_str(x) -> str:
    return ("" if x is None else str(x)).strip()

def _safe_lower(x) -> str:
    return _norm_str(x).lower()

def _strip_accents(s: str) -> str:
    """
    Remove acentos para facilitar match de cabeçalhos e geração de chave.
    """
    s = _norm_str(s)
    if not s:
        return ""
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", s)
        if not unicodedata.combining(ch)
    )

def _norm_header_name(s: str) -> str:
    """
    Normaliza cabeçalhos:
    - lower
    - sem acento
    - remove pontuação básica
    - colapsa espaços
    """
    s = _strip_accents(s).lower()
    s = s.replace("_", " ")
    s = re.sub(r"[^\w\s]", " ", s)     # tira pontuação
    s = re.sub(r"\s+", " ", s).strip() # colapsa espaços
    return s

def _key_norm(s: str) -> str:
    """
    Normalização para montar a chave única.
    """
    s = _strip_accents(s).lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _make_chave(nome: str, email: str, cpf: str, whatsapp: str) -> str:
    """
    Chave única por criança + contato do responsável.
    - se tiver email, usa email
    - senão, usa cpf
    - senão, usa whatsapp
    - senão, "sem-contato"
    """
    nome_n = _key_norm(nome)
    contato = _key_norm(email) or _key_norm(cpf) or _key_norm(whatsapp) or "sem-contato"
    return f"{nome_n}|{contato}"

def _pg_unique_msg(exc: Exception) -> str:
    """
    Tenta traduzir violação de UNIQUE no Postgres/pg8000.
    """
    msg = (str(exc) or "").lower()
    if "unique" in msg or "duplicate key" in msg:
        if "chave" in msg:
            return "Registro duplicado (mesma criança e mesmo contato)."
        return "Registro duplicado (chave única)."
    return "Erro ao salvar dados."


# -----------------------------
# CADASTRAR / ATUALIZAR (UPSERT) - AGORA POR CHAVE
# -----------------------------

def cadastrar_participante(nome, email, cpf, whatsapp, curso, perfil, semestre):
    """
    Regras (Opção B):
    - nome obrigatório
    - email/cpf NÃO são mais únicos (mãe pode ter 2 filhos)
    - UPSERT por 'chave' (nome da criança + contato)
    """
    nome = _norm_str(nome)
    email = _norm_str(email)
    cpf = _norm_str(cpf)
    whatsapp = _norm_str(whatsapp)
    curso = _norm_str(curso)
    perfil = _norm_str(perfil)
    semestre = _norm_str(semestre)

    if not nome:
        raise ValueError("Nome é obrigatório.")

    chave = _make_chave(nome, email, cpf, whatsapp)

    conn = conectar()
    cur = conn.cursor()

    try:
        cur.execute(
            """
            INSERT INTO participantes
                (chave, nome, email, cpf, whatsapp, curso, perfil, semestre, status, bloqueado, confirmado)
            VALUES
                (%s, %s, %s, %s, %s, %s, %s, %s, 'INSCRITO', FALSE, FALSE)
            ON CONFLICT (chave) DO UPDATE SET
                nome       = EXCLUDED.nome,
                email      = EXCLUDED.email,
                cpf        = EXCLUDED.cpf,
                whatsapp   = EXCLUDED.whatsapp,
                curso      = EXCLUDED.curso,
                perfil     = EXCLUDED.perfil,
                semestre   = EXCLUDED.semestre,
                status     = 'INSCRITO',
                bloqueado  = FALSE,
                confirmado = FALSE
            """,
            (chave, nome, email or None, cpf or None, whatsapp, curso, perfil, semestre),
        )

        conn.commit()

    except Exception as e:
        conn.rollback()
        raise ValueError(_pg_unique_msg(e))
    finally:
        conn.close()


# -----------------------------
# CONFIRMAR PRESENÇA
# -----------------------------

def confirmar_presenca_por_email(email: str) -> int:
    conn = conectar()
    cur = conn.cursor()

    email_norm = _safe_lower(email)
    if not email_norm:
        conn.close()
        return 0

    cur.execute(
        """
        UPDATE participantes
        SET confirmado = TRUE,
            status = 'CONFIRMADO'
        WHERE status = 'SELECIONADO'
          AND email IS NOT NULL
          AND LOWER(TRIM(email)) = %s
        """,
        (email_norm,),
    )
    alterados = cur.rowcount
    conn.commit()
    conn.close()
    return alterados


# -----------------------------
# PROMOVER SUPLENTE
# -----------------------------

def promover_suplente_se_expirou(prazo_horas: int = 48) -> Tuple[bool, str]:
    if prazo_horas < 0:
        return False, "prazo_horas inválido (deve ser >= 0)."

    limite_dt = datetime.now() - timedelta(hours=prazo_horas)

    conn = conectar()
    cur = conn.cursor()

    try:
        # 1) pega 1 selecionado expirado (não confirmado)
        cur.execute(
            """
            SELECT id, prioridade
            FROM participantes
            WHERE status='SELECIONADO'
              AND confirmado=FALSE
              AND data_sorteio IS NOT NULL
              AND data_sorteio <= %s
            ORDER BY data_sorteio ASC
            LIMIT 1
            """,
            (limite_dt,),
        )
        expirado = cur.fetchone()
        if not expirado:
            return False, "Nenhum selecionado expirado encontrado."

        expirado_id = expirado[0]
        expirado_prioridade = expirado[1]

        # 2) pega o primeiro suplente
        cur.execute(
            """
            SELECT id, nome, email
            FROM participantes
            WHERE status='SUPLENTE'
            ORDER BY prioridade ASC
            LIMIT 1
            """
        )
        suplente = cur.fetchone()

        if not suplente:
            cur.execute(
                """
                UPDATE participantes
                SET status='EXPIRADO', bloqueado=FALSE
                WHERE id=%s
                """,
                (expirado_id,),
            )
            conn.commit()
            return False, "Nenhum suplente disponível; selecionado marcado como EXPIRADO."

        supl_id, supl_nome, supl_email = suplente[0], (suplente[1] or ""), (suplente[2] or "")
        agora = datetime.now()

        # 3) promove suplente
        cur.execute(
            """
            UPDATE participantes
            SET status='SELECIONADO',
                bloqueado=TRUE,
                confirmado=FALSE,
                data_sorteio=%s,
                prioridade=%s
            WHERE id=%s
            """,
            (agora, expirado_prioridade, supl_id),
        )

        # 4) marca expirado
        cur.execute(
            """
            UPDATE participantes
            SET status='EXPIRADO', bloqueado=FALSE
            WHERE id=%s
            """,
            (expirado_id,),
        )

        # 5) histórico
        registrar_vencedor(supl_id, supl_nome, supl_email, conn=conn)

        conn.commit()
        return True, f"Suplente promovido: {supl_nome} ({supl_email})."

    except Exception as e:
        conn.rollback()
        return False, f"Erro: {type(e).__name__}: {e}"
    finally:
        conn.close()


# -----------------------------
# HISTÓRICO
# -----------------------------

def listar_historico() -> List[Tuple[str, str, str, str]]:
    conn = conectar()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT h.data_sorteio, h.nome, h.email, COALESCE(p.status, 'REMOVIDO')
        FROM historico_sorteios h
        LEFT JOIN participantes p ON p.id = h.participante_id
        ORDER BY h.id DESC
        """
    )
    rows = cur.fetchall()
    conn.close()
    return [(r[0], r[1], r[2], r[3]) for r in rows]


def limpar_historico(reset_id: bool = True) -> int:
    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM historico_sorteios")
    total = int(cur.fetchone()[0] or 0)

    if reset_id:
        cur.execute("TRUNCATE TABLE historico_sorteios RESTART IDENTITY")
    else:
        cur.execute("DELETE FROM historico_sorteios")

    conn.commit()
    conn.close()
    return total


# -----------------------------
# EXPORTAR XLSX
# -----------------------------

def exportar_participantes_xlsx() -> bytes:
    conn = conectar()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, nome, email, cpf, whatsapp, curso, perfil, semestre,
               status, confirmado, data_sorteio, prioridade
        FROM participantes
        ORDER BY id ASC
        """
    )
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Participantes"

    ws.append([
        "id", "nome", "email", "cpf", "whatsapp", "curso", "perfil",
        "semestre", "status", "confirmado", "data_sorteio", "prioridade"
    ])

    for r in rows:
        ws.append(list(r))

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -----------------------------
# IMPORTAR XLSX
# -----------------------------

def importar_participantes_xlsx(file_bytes: bytes) -> Dict[str, Any]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Cabeçalho normalizado (com remoção de acentos e pontuação)
    header_raw = []
    header_norm = []
    for cell in ws[1]:
        v = "" if cell.value is None else str(cell.value)
        header_raw.append(v)
        header_norm.append(_norm_header_name(v))

    def idx(*names: str) -> Optional[int]:
        for n in names:
            nn = _norm_header_name(n)
            if nn in header_norm:
                return header_norm.index(nn)
        return None

    col_nome = idx(
        "nome",
        "nome da crianca",
        "nome da criança",
        "nome completo",
        "crianca",
        "criança",
    )

    col_email = idx(
        "email", "e-mail", "e_mail",
        "e mail para contato",
        "email para contato",
        "e-mail para contato",
    )

    col_whats = idx(
        "whatsapp", "whats", "telefone", "celular",
        "telefone de contato do(a) responsavel",
        "telefone de contato do(a) responsável",
        "telefone de contato do responsavel",
        "telefone de contato do responsável",
    )

    col_cpf = idx(
        "cpf",
        "documento",
        "documento do(a) responsavel e tipo de documento",
        "documento do(a) responsável e tipo de documento",
        "rg",
        "cpf/rg",
    )

    col_curso = idx("curso")
    col_perfil = idx("perfil")
    col_semestre = idx("semestre")

    if col_nome is None:
        return {
            "ok": False,
            "msg": "Planilha inválida: não encontrei uma coluna de NOME (ex: 'Nome da criança' ou 'nome').",
            "importados": 0,
            "erros": [],
            "cabecalhos_detectados": header_raw[:],
        }

    importados = 0
    erros: List[Dict[str, Any]] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            nome = _norm_str(row[col_nome]) if col_nome is not None else ""
            email = _norm_str(row[col_email]) if col_email is not None else ""
            whatsapp = _norm_str(row[col_whats]) if col_whats is not None else ""
            cpf = _norm_str(row[col_cpf]) if col_cpf is not None else ""

            curso = _norm_str(row[col_curso]) if col_curso is not None else ""
            perfil = _norm_str(row[col_perfil]) if col_perfil is not None else ""
            semestre = _norm_str(row[col_semestre]) if col_semestre is not None else ""

            if not nome and not email and not cpf:
                continue

            cadastrar_participante(
                nome=nome,
                email=email,
                cpf=cpf,
                whatsapp=whatsapp,
                curso=curso,
                perfil=perfil,
                semestre=semestre
            )
            importados += 1

        except Exception as e:
            erros.append({"linha": i, "erro": str(e)})

    return {
        "ok": True,
        "msg": f"Importação finalizada. {importados} linha(s) processada(s).",
        "importados": importados,
        "erros": erros,
        "erros_qtd": len(erros),
    }
